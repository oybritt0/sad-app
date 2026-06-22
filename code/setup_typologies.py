"""
setup_typologies.py

Ingest the SAD_Typologies.xlsx classification spreadsheet into the
pipeline's filesystem as canonical JSON, ready for any module to
consume without needing to read the spreadsheet directly.

WHAT THIS DOES
    Reads two sheets from the workbook:
        - 'in'                — sad_id, sad_name, anchor_venue,
                                PrimaryTypology, notes (per-SAD source rows)
        - 'PrimarySecondary'  — sad_name, anchor_venue, Primary/Secondary
                                Typology, Notes, Links (richer)
    Matches them by `sad_name` (stripped) and merges into one canonical
    record per SAD.

OUTPUTS
    <data-dir>/_shared/typologies.json
        Canonical, keyed by sad_id. Single source of truth — modules
        should read this.

    <data-dir>/<sad_id>/derived/typology.json
        Per-SAD copy for the modules that already operate on
        derived/ folders. Identical content to one entry in the
        canonical file.

USAGE
    python setup_typologies.py \\
        --xlsx "<path-to-SAD_Typologies.xlsx>" \\
        --data-dir "<path-to-data-folder>"

    Idempotent — safe to re-run. Overwrites existing typology.json
    files. Run after adjusting the spreadsheet to push updates
    downstream.

SCHEMA OF EACH RECORD
    {
        "sad_id":             "08_Water-Street_Tampa-FL",
        "sad_name":           "Water Street, Tampa-FL",
        "anchor_venue":       "Amalie Arena",
        "primary_typology":   "Innovation",
        "secondary_typology": "Community",
        "notes":              null,
        "links":              null,
        "source_sheets":      ["in", "PrimarySecondary"]
    }
"""
from __future__ import annotations
import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl",
          file=sys.stderr)
    sys.exit(1)


# Known typology values (for validation / catching new ones)
KNOWN_PRIMARY_TYPOLOGIES = {
    'Entertainment', 'Community', 'Innovation', 'Sports Park',
}


def _norm(s):
    """Normalize a cell value for matching: strip whitespace, None-safe."""
    if s is None:
        return None
    s = str(s).strip()
    return s if s else None


def read_in_sheet(wb) -> dict:
    """Read the 'in' sheet. Returns {sad_id: record_dict}."""
    ws = wb['in']
    hdr = [c.value for c in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rec = dict(zip(hdr, row))
        sad_id = _norm(rec.get('sad_id'))
        if not sad_id:
            continue
        out[sad_id] = {
            'sad_id':           sad_id,
            'sad_name':         _norm(rec.get('sad_name')),
            'anchor_venue':     _norm(rec.get('anchor_venue')),
            'primary_typology': _norm(rec.get('PrimaryTypology')),
            'notes':            _norm(rec.get('notes')),
        }
    return out


def read_primarysecondary_sheet(wb) -> dict:
    """Read 'PrimarySecondary'. Returns {sad_name: record_dict}.

    Skips methodology-note rows that lack a primary typology.
    """
    ws = wb['PrimarySecondary']
    hdr = [c.value for c in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rec = dict(zip(hdr, row))
        primary = _norm(rec.get('Primary Typology'))
        if not primary:
            # Methodology-note row, not a SAD row
            continue
        sad_name = _norm(rec.get('SAD_name'))
        if not sad_name:
            continue
        out[sad_name] = {
            'sad_name':           sad_name,
            'anchor_venue':       _norm(rec.get('anchor_venue')),
            'primary_typology':   primary,
            'secondary_typology': _norm(rec.get('Secondary Typology')),
            'notes':              _norm(rec.get('Notes')),
            'links':              _norm(rec.get('Links')),
        }
    return out


def merge_records(in_by_id: dict, ps_by_name: dict) -> dict:
    """Merge the two sheets into one record per sad_id."""
    merged = {}
    for sad_id, in_rec in in_by_id.items():
        sname = in_rec.get('sad_name')
        ps_rec = ps_by_name.get(sname) if sname else None

        # Primary: prefer PrimarySecondary (richer), fall back to 'in'
        primary = (ps_rec or {}).get('primary_typology') \
                  or in_rec.get('primary_typology')

        # Notes: prefer the longer/non-empty
        notes_a = in_rec.get('notes')
        notes_b = (ps_rec or {}).get('notes')
        notes = notes_b or notes_a

        merged[sad_id] = {
            'sad_id':             sad_id,
            'sad_name':           in_rec.get('sad_name'),
            'anchor_venue':       in_rec.get('anchor_venue')
                                  or (ps_rec or {}).get('anchor_venue'),
            'primary_typology':   primary,
            'secondary_typology': (ps_rec or {}).get('secondary_typology'),
            'notes':              notes,
            'links':              (ps_rec or {}).get('links'),
            'source_sheets':      ['in'] + (['PrimarySecondary'] if ps_rec else []),
        }
    return merged


def validate(merged: dict) -> list:
    """Return a list of warning strings (empty if all clean)."""
    warnings = []
    seen_primary = set()
    seen_secondary = set()
    for sad_id, rec in merged.items():
        if not rec.get('primary_typology'):
            warnings.append(f"  {sad_id}: missing primary_typology")
        elif rec['primary_typology'] not in KNOWN_PRIMARY_TYPOLOGIES:
            warnings.append(
                f"  {sad_id}: unrecognized primary typology "
                f"{rec['primary_typology']!r} — add to "
                f"KNOWN_PRIMARY_TYPOLOGIES if intentional")
        if rec.get('primary_typology'):
            seen_primary.add(rec['primary_typology'])
        if rec.get('secondary_typology'):
            seen_secondary.add(rec['secondary_typology'])
    if not warnings:
        # Add an informational note about what was seen
        warnings = None
    return warnings, seen_primary, seen_secondary


def write_canonical(merged: dict, data_dir: Path):
    """Write the master typologies.json at <data-dir>/_shared/."""
    shared_dir = data_dir / '_shared'
    shared_dir.mkdir(parents=True, exist_ok=True)
    out_path = shared_dir / 'typologies.json'

    payload = {
        'description': ('Per-SAD typology classification from '
                        'SAD_Typologies.xlsx. Use sad_id as the lookup key. '
                        'Generated by setup_typologies.py — re-run after '
                        'editing the source spreadsheet.'),
        'typologies':       sorted(KNOWN_PRIMARY_TYPOLOGIES),
        'count':            len(merged),
        'districts':        merged,
    }
    out_path.write_text(json.dumps(payload, indent=2))
    return out_path


def write_per_sad(merged: dict, data_dir: Path) -> tuple:
    """Write one typology.json per SAD to <data-dir>/<sad_id>/derived/."""
    written, missing_derived = 0, []
    for sad_id, rec in merged.items():
        derived = data_dir / sad_id / 'derived'
        if not derived.exists():
            missing_derived.append(sad_id)
            continue
        (derived / 'typology.json').write_text(json.dumps(rec, indent=2))
        written += 1
    return written, missing_derived


def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--xlsx', type=Path, required=True,
                    help='Path to SAD_Typologies.xlsx')
    ap.add_argument('--data-dir', type=Path, required=True,
                    help='Pipeline data root (the same path used by '
                         'batch_run_pipeline.py)')
    ap.add_argument('--skip-per-sad', action='store_true',
                    help='Only write the canonical _shared/typologies.json '
                         'and skip writing per-SAD derived/typology.json '
                         'files.')
    args = ap.parse_args()

    if not args.xlsx.exists():
        sys.exit(f"ERROR: spreadsheet not found at {args.xlsx}")
    if not args.data_dir.exists():
        sys.exit(f"ERROR: data-dir not found at {args.data_dir}")

    print(f"Loading {args.xlsx.name}...")
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)

    in_by_id = read_in_sheet(wb)
    ps_by_name = read_primarysecondary_sheet(wb)
    print(f"  'in' sheet:               {len(in_by_id)} SADs")
    print(f"  'PrimarySecondary' sheet: {len(ps_by_name)} SADs")

    merged = merge_records(in_by_id, ps_by_name)
    print(f"  merged:                   {len(merged)} SADs")

    # Validation
    warnings, primary_seen, secondary_seen = validate(merged)
    if warnings:
        print("\nWarnings:")
        for w in warnings:
            print(w)

    # Typology distribution
    from collections import Counter
    primary_counts = Counter(rec.get('primary_typology') for rec in merged.values())
    secondary_counts = Counter(rec.get('secondary_typology')
                               for rec in merged.values()
                               if rec.get('secondary_typology'))
    print("\nPrimary typology distribution:")
    for t, n in primary_counts.most_common():
        print(f"  {(t or 'UNCLASSIFIED'):20s} {n}")
    print("\nSecondary typology distribution (where present):")
    for t, n in secondary_counts.most_common():
        print(f"  {t:20s} {n}")

    # Write canonical
    canonical_path = write_canonical(merged, args.data_dir)
    print(f"\nWrote canonical: {canonical_path}")

    # Write per-SAD
    if not args.skip_per_sad:
        written, missing = write_per_sad(merged, args.data_dir)
        print(f"Wrote per-SAD typology.json to {written} districts")
        if missing:
            print(f"  Skipped {len(missing)} SADs (no derived/ folder yet):")
            for sad_id in missing[:5]:
                print(f"    {sad_id}")
            if len(missing) > 5:
                print(f"    ... and {len(missing) - 5} more")
    else:
        print("Skipped per-SAD writes (--skip-per-sad)")

    print("\nDone.")


if __name__ == '__main__':
    main()
